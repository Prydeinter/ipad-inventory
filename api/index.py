"""Vercel Serverless entry point untuk AAIIBS iPad Lungsuran System.
Semua route /api/* di-route ke sini via vercel.json.
"""
import os
import io
import logging
import secrets
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, UploadFile, File
from fastapi.responses import Response, JSONResponse
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr
import uuid

# ---------------------------------------------------------------- setup
MONGO_URL = os.environ.get("MONGO_URL", "")
DB_NAME = os.environ.get("DB_NAME", "aaiibs_ipad")
JWT_SECRET = os.environ.get("JWT_SECRET", "dev-secret-change-in-production")
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@aaiibs.sch.id").lower()
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "AAIIBS@2026")
CORS_ORIGINS = os.environ.get("CORS_ORIGINS", "*").split(",")

JWT_ALG = "HS256"
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("aaiibs")

app = FastAPI(title="AAIIBS iPad Lungsuran System")
api = APIRouter(prefix="/api")

# ---------------------------------------------------------------- lazy Mongo client
_client = None
_db = None
_seeded = False


def get_db():
    """Lazy init MongoDB client (dipanggil hanya saat endpoint dieksekusi)."""
    global _client, _db
    if _db is None:
        from motor.motor_asyncio import AsyncIOMotorClient
        _client = AsyncIOMotorClient(MONGO_URL, serverSelectionTimeoutMS=10000)
        _db = _client[DB_NAME]
    return _db


# ---------------------------------------------------------------- auth utils
def hash_password(pw: str) -> str:
    import bcrypt
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        import bcrypt
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_access_token(uid: str, email: str) -> str:
    import jwt
    payload = {"sub": uid, "email": email, "type": "access",
               "exp": datetime.now(timezone.utc) + timedelta(hours=12)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def ensure_seed():
    """Idempotent seed admin + indexes."""
    global _seeded
    if _seeded:
        return
    try:
        db = get_db()
        await db.users.create_index("email", unique=True)
        await db.ipads.create_index("serial_number", unique=True)
        await db.codes.create_index("code", unique=True)
        await db.paktas.create_index("serial_number")
        existing = await db.users.find_one({"email": ADMIN_EMAIL})
        if not existing:
            await db.users.insert_one({
                "email": ADMIN_EMAIL,
                "password_hash": hash_password(ADMIN_PASSWORD),
                "name": "Administrator",
                "role": "admin",
                "created_at": now_iso(),
            })
            logger.info("Seeded admin user %s", ADMIN_EMAIL)
        elif not verify_password(ADMIN_PASSWORD, existing["password_hash"]):
            await db.users.update_one({"email": ADMIN_EMAIL},
                                      {"$set": {"password_hash": hash_password(ADMIN_PASSWORD)}})
        _seeded = True
    except Exception as e:
        logger.exception("Seed error: %s", e)


async def get_current_admin(request: Request) -> dict:
    import jwt
    from bson import ObjectId
    token = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
    if not token:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        db = get_db()
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return {"id": str(user["_id"]), "email": user["email"], "name": user.get("name", "Admin")}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ---------------------------------------------------------------- models
class LoginInput(BaseModel):
    email: EmailStr
    password: str


class IpadInput(BaseModel):
    serial_number: str
    version: str
    storage: str
    purchase_year: int
    color: Optional[str] = ""
    notes: Optional[str] = ""


class CodeInput(BaseModel):
    serial_number: str
    target_name: Optional[str] = ""
    count: int = 1


class BatchCodeInput(BaseModel):
    serial_numbers: List[str]
    target_name: Optional[str] = ""
    count: int = 1


class ValidateInput(BaseModel):
    code: str


class PaktaInput(BaseModel):
    code: str
    nama: str
    nik: str
    jabatan: str
    unit: str
    tanggal_peminjaman: str
    signature: str


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def gen_code(n=6):
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(secrets.choice(alphabet) for _ in range(n))


# ---------------------------------------------------------------- global exception handler
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.exception("Unhandled error on %s: %s", request.url.path, exc)
    return JSONResponse(status_code=500,
                        content={"detail": f"Internal error: {type(exc).__name__}: {str(exc)[:200]}"})


# ---------------------------------------------------------------- health / debug
@api.get("/")
async def root():
    return {"message": "AAIIBS iPad Lungsuran System API", "status": "ok"}


@api.get("/health")
async def health():
    """Health check tanpa touch DB."""
    return {
        "status": "ok",
        "env": {
            "MONGO_URL": "SET" if MONGO_URL else "MISSING",
            "DB_NAME": DB_NAME,
            "JWT_SECRET": "SET" if JWT_SECRET else "MISSING",
            "ADMIN_EMAIL": "SET" if ADMIN_EMAIL else "MISSING",
            "ADMIN_PASSWORD": "SET" if ADMIN_PASSWORD else "MISSING",
        }
    }


@api.get("/health/db")
async def health_db():
    """Health check dengan ping MongoDB."""
    try:
        db = get_db()
        await db.command("ping")
        return {"status": "ok", "db": DB_NAME, "mongodb": "connected"}
    except Exception as e:
        return JSONResponse(status_code=500,
                            content={"status": "error",
                                     "detail": f"{type(e).__name__}: {str(e)[:300]}"})


# ---------------------------------------------------------------- auth routes
@api.post("/auth/login")
async def login(body: LoginInput, response: Response):
    await ensure_seed()
    db = get_db()
    email = body.email.lower()
    ident = f"login:{email}"
    att = await db.login_attempts.find_one({"identifier": ident})
    if att and att.get("count", 0) >= 5:
        locked_until = att.get("locked_until")
        if locked_until and datetime.fromisoformat(locked_until) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Terlalu banyak percobaan. Coba lagi nanti.")
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": ident},
            {"$inc": {"count": 1},
             "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}},
            upsert=True,
        )
        raise HTTPException(status_code=401, detail="Email atau password salah")
    await db.login_attempts.delete_one({"identifier": ident})
    token = create_access_token(str(user["_id"]), user["email"])
    response.set_cookie("access_token", token, httponly=True, secure=True,
                        samesite="lax", max_age=43200, path="/")
    return {"access_token": token,
            "user": {"id": str(user["_id"]), "email": user["email"], "name": user.get("name", "Admin")}}


@api.get("/auth/me")
async def me(admin: dict = Depends(get_current_admin)):
    return admin


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


# ---------------------------------------------------------------- ipad admin routes
@api.post("/admin/ipads")
async def add_ipad(body: IpadInput, admin: dict = Depends(get_current_admin)):
    db = get_db()
    serial = body.serial_number.strip().upper()
    if await db.ipads.find_one({"serial_number": serial}):
        raise HTTPException(status_code=400, detail="Serial number sudah terdaftar")
    doc = {
        "id": str(uuid.uuid4()),
        "serial_number": serial,
        "version": body.version.strip(),
        "storage": body.storage.strip(),
        "purchase_year": body.purchase_year,
        "color": (body.color or "").strip(),
        "notes": (body.notes or "").strip(),
        "created_at": now_iso(),
    }
    await db.ipads.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/admin/ipads/{ipad_id}")
async def update_ipad(ipad_id: str, body: IpadInput, admin: dict = Depends(get_current_admin)):
    db = get_db()
    serial = body.serial_number.strip().upper()
    existing = await db.ipads.find_one({"id": ipad_id})
    if not existing:
        raise HTTPException(status_code=404, detail="iPad tidak ditemukan")
    clash = await db.ipads.find_one({"serial_number": serial, "id": {"$ne": ipad_id}})
    if clash:
        raise HTTPException(status_code=400, detail="Serial number sudah dipakai iPad lain")
    await db.ipads.update_one({"id": ipad_id}, {"$set": {
        "serial_number": serial, "version": body.version.strip(), "storage": body.storage.strip(),
        "purchase_year": body.purchase_year, "color": (body.color or "").strip(),
        "notes": (body.notes or "").strip(),
    }})
    return {"ok": True}


@api.delete("/admin/ipads/{ipad_id}")
async def delete_ipad(ipad_id: str, admin: dict = Depends(get_current_admin)):
    db = get_db()
    ipad = await db.ipads.find_one({"id": ipad_id})
    if not ipad:
        raise HTTPException(status_code=404, detail="iPad tidak ditemukan")
    if await db.paktas.find_one({"serial_number": ipad["serial_number"]}):
        raise HTTPException(status_code=400, detail="Tidak bisa dihapus, sudah ada riwayat pakta.")
    await db.ipads.delete_one({"id": ipad_id})
    await db.codes.delete_many({"serial_number": ipad["serial_number"], "status": "active"})
    return {"ok": True}


@api.get("/admin/ipads")
async def admin_list_ipads(admin: dict = Depends(get_current_admin)):
    return await _ipads_with_holders()


IMPORT_HEADERS = ["serial_number", "version", "storage", "purchase_year", "color", "notes"]


@api.get("/admin/ipads/template")
async def ipad_template(admin: dict = Depends(get_current_admin)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "iPad"
    ws.append(IMPORT_HEADERS)
    ws.append(["DMPGX1AAAA01", "iPad Gen 10", "256GB", 2026, "Silver", "unit baru"])
    ws.append(["F9FZ2BBBB03", "iPad Gen 7", "32GB", 2021, "Space Gray", "lungsuran"])
    for i, col in enumerate(["A", "B", "C", "D", "E", "F"]):
        ws.column_dimensions[col].width = [18, 16, 12, 14, 14, 24][i]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_import_ipad.xlsx"'},
    )


@api.post("/admin/ipads/bulk")
async def bulk_import_ipads(file: UploadFile = File(...), admin: dict = Depends(get_current_admin)):
    from openpyxl import load_workbook
    db = get_db()
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File harus berformat Excel (.xlsx)")
    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Gagal membaca file Excel")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    created, skipped, errors = 0, 0, []
    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = list(row) + [None] * (6 - len(row))
        serial = str(cells[0]).strip().upper() if cells[0] else ""
        version = str(cells[1]).strip() if cells[1] else ""
        storage = str(cells[2]).strip() if cells[2] else ""
        try:
            year = int(float(cells[3])) if cells[3] not in (None, "") else 0
        except (ValueError, TypeError):
            year = 0
        color = str(cells[4]).strip() if cells[4] else ""
        notes = str(cells[5]).strip() if cells[5] else ""
        if not serial or not version or not storage or not year:
            errors.append(f"Baris {idx}: serial/versi/penyimpanan/tahun wajib diisi")
            continue
        if await db.ipads.find_one({"serial_number": serial}):
            skipped += 1
            continue
        await db.ipads.insert_one({
            "id": str(uuid.uuid4()), "serial_number": serial, "version": version,
            "storage": storage, "purchase_year": year, "color": color,
            "notes": notes, "created_at": now_iso(),
        })
        created += 1
    return {"created": created, "skipped": skipped, "errors": errors[:20]}


# ---------------------------------------------------------------- code admin routes
@api.post("/admin/codes")
async def create_codes(body: CodeInput, admin: dict = Depends(get_current_admin)):
    db = get_db()
    ipad = await db.ipads.find_one({"serial_number": body.serial_number.strip().upper()})
    if not ipad:
        raise HTTPException(status_code=404, detail="iPad tidak ditemukan")
    return await _gen_codes_for_ipad(ipad, body.count, body.target_name)


async def _gen_codes_for_ipad(ipad, count, target_name):
    db = get_db()
    count = max(1, min(int(count), 20))
    created = []
    for _ in range(count):
        code = gen_code()
        while await db.codes.find_one({"code": code}):
            code = gen_code()
        doc = {
            "id": str(uuid.uuid4()), "code": code,
            "serial_number": ipad["serial_number"], "version": ipad["version"],
            "storage": ipad["storage"], "purchase_year": ipad["purchase_year"],
            "target_name": (target_name or "").strip(),
            "status": "active", "used_by": None,
            "created_at": now_iso(), "used_at": None,
        }
        await db.codes.insert_one(doc)
        doc.pop("_id", None)
        created.append(doc)
    return created


@api.post("/admin/codes/batch")
async def create_codes_batch(body: BatchCodeInput, admin: dict = Depends(get_current_admin)):
    db = get_db()
    serials = list({s.strip().upper() for s in body.serial_numbers if s and s.strip()})
    if not serials:
        raise HTTPException(status_code=400, detail="Pilih minimal satu iPad")
    created, not_found = [], []
    for serial in serials:
        ipad = await db.ipads.find_one({"serial_number": serial})
        if not ipad:
            not_found.append(serial)
            continue
        created.extend(await _gen_codes_for_ipad(ipad, body.count, body.target_name))
    return {"created": created, "ipad_count": len(serials) - len(not_found),
            "code_count": len(created), "not_found": not_found}


@api.get("/admin/codes")
async def admin_list_codes(admin: dict = Depends(get_current_admin)):
    db = get_db()
    codes = await db.codes.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return codes


@api.delete("/admin/codes/{code_id}")
async def delete_code(code_id: str, admin: dict = Depends(get_current_admin)):
    db = get_db()
    code = await db.codes.find_one({"id": code_id})
    if not code:
        raise HTTPException(status_code=404, detail="Kode tidak ditemukan")
    if code["status"] == "used":
        raise HTTPException(status_code=400, detail="Kode sudah dipakai, tidak bisa dihapus")
    await db.codes.delete_one({"id": code_id})
    return {"ok": True}


@api.get("/admin/paktas")
async def admin_list_paktas(admin: dict = Depends(get_current_admin)):
    db = get_db()
    return await db.paktas.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


# ---------------------------------------------------------------- public: token + pakta
@api.post("/codes/validate")
async def validate_code(body: ValidateInput):
    db = get_db()
    code = body.code.strip().upper()
    doc = await db.codes.find_one({"code": code}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Kode tidak valid")
    if doc["status"] == "used":
        raise HTTPException(status_code=400, detail="Kode ini sudah digunakan untuk mengisi pakta")
    return {
        "code": doc["code"], "serial_number": doc["serial_number"],
        "version": doc["version"], "storage": doc["storage"],
        "purchase_year": doc["purchase_year"], "target_name": doc.get("target_name", ""),
    }


@api.post("/pakta/submit")
async def submit_pakta(body: PaktaInput):
    db = get_db()
    code = body.code.strip().upper()
    codedoc = await db.codes.find_one({"code": code})
    if not codedoc:
        raise HTTPException(status_code=404, detail="Kode tidak valid")
    if codedoc["status"] == "used":
        raise HTTPException(status_code=400, detail="Kode ini sudah digunakan")
    if not body.signature or not body.signature.startswith("data:image") or "," not in body.signature:
        raise HTTPException(status_code=400, detail="Tanda tangan wajib diisi")

    prior = await db.paktas.count_documents({"serial_number": codedoc["serial_number"]})
    pakta = {
        "id": str(uuid.uuid4()),
        "serial_number": codedoc["serial_number"],
        "ipad_version": codedoc["version"],
        "storage": codedoc["storage"],
        "purchase_year": codedoc["purchase_year"],
        "nama": body.nama.strip(),
        "nik": body.nik.strip(),
        "jabatan": body.jabatan.strip(),
        "unit": body.unit.strip(),
        "tanggal_peminjaman": body.tanggal_peminjaman,
        "tanggal_pengisian": now_iso(),
        "signature": body.signature,
        "nama_terang": body.nama.strip(),
        "code": code,
        "sequence": prior + 1,
        "created_at": now_iso(),
    }
    await db.paktas.insert_one(pakta)
    await db.codes.update_one({"code": code},
                              {"$set": {"status": "used", "used_by": pakta["id"], "used_at": now_iso()}})
    return {"id": pakta["id"]}


@api.get("/pakta/{pakta_id}")
async def get_pakta(pakta_id: str):
    db = get_db()
    doc = await db.paktas.find_one({"id": pakta_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Pakta tidak ditemukan")
    return doc


@api.get("/pakta/{pakta_id}/pdf")
async def pakta_pdf(pakta_id: str):
    db = get_db()
    doc = await db.paktas.find_one({"id": pakta_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Pakta tidak ditemukan")
    # lazy import biar reportlab hanya di-load saat generate PDF
    from pdf_pakta import build_pakta_pdf
    pdf = build_pakta_pdf(doc)
    safe = doc["nama"].replace(" ", "_")
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="Pakta_{safe}.pdf"'})


# ---------------------------------------------------------------- public dashboard / trace
async def _ipads_with_holders():
    db = get_db()
    ipads = await db.ipads.find({}, {"_id": 0}).sort("purchase_year", -1).to_list(1000)
    out = []
    for ip in ipads:
        chain = await db.paktas.find(
            {"serial_number": ip["serial_number"]}, {"_id": 0, "signature": 0}
        ).sort("created_at", 1).to_list(100)
        current = chain[-1] if chain else None
        out.append({
            **ip,
            "current_holder": current["nama"] if current else None,
            "current_holder_unit": (f"{current['jabatan']} / {current['unit']}" if current else None),
            "holder_count": len(chain),
            "is_lungsuran": len(chain) > 1,
            "chain": chain,
        })
    return out


@api.get("/public/ipads")
async def public_ipads():
    return await _ipads_with_holders()


@api.get("/public/trace/{serial}")
async def public_trace(serial: str):
    db = get_db()
    serial = serial.strip().upper()
    ipad = await db.ipads.find_one({"serial_number": serial}, {"_id": 0})
    if not ipad:
        raise HTTPException(status_code=404, detail="Serial number tidak ditemukan")
    chain = await db.paktas.find(
        {"serial_number": serial}, {"_id": 0, "signature": 0}
    ).sort("created_at", 1).to_list(100)
    return {"ipad": ipad, "chain": chain}


@api.get("/public/stats")
async def public_stats():
    db = get_db()
    ipads = await db.ipads.find({}, {"_id": 0}).to_list(1000)
    paktas = await db.paktas.find({}, {"_id": 0, "signature": 0}).to_list(2000)

    by_version, by_storage, by_year = {}, {}, {}
    for ip in ipads:
        by_version[ip["version"]] = by_version.get(ip["version"], 0) + 1
        by_storage[ip["storage"]] = by_storage.get(ip["storage"], 0) + 1
        y = str(ip["purchase_year"])
        by_year[y] = by_year.get(y, 0) + 1

    serials_with_holder = {p["serial_number"] for p in paktas}
    chain_counts = {}
    for p in paktas:
        chain_counts[p["serial_number"]] = chain_counts.get(p["serial_number"], 0) + 1
    lungsuran_count = sum(1 for v in chain_counts.values() if v > 1)

    recent = sorted(paktas, key=lambda x: x.get("created_at", ""), reverse=True)[:8]
    recent = [{"nama": r["nama"], "unit": r["unit"], "jabatan": r["jabatan"],
               "serial_number": r["serial_number"], "ipad_version": r["ipad_version"],
               "storage": r["storage"], "created_at": r["created_at"], "id": r["id"]} for r in recent]

    active_codes = await db.codes.count_documents({"status": "active"})

    return {
        "total_ipads": len(ipads),
        "total_paktas": len(paktas),
        "active_holders": len(serials_with_holder),
        "available_ipads": len(ipads) - len(serials_with_holder),
        "lungsuran_count": lungsuran_count,
        "active_codes": active_codes,
        "by_version": [{"name": k, "value": v} for k, v in sorted(by_version.items())],
        "by_storage": [{"name": k, "value": v} for k, v in sorted(by_storage.items())],
        "by_year": [{"name": k, "value": v} for k, v in sorted(by_year.items())],
        "recent": recent,
    }


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=CORS_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)
