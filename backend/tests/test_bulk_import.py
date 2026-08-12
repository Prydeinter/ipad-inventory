"""Tests for new bulk import + template download endpoints."""
import io
import os
import requests
import pytest
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"
ADMIN_EMAIL = "admin@aaiibs.sch.id"
ADMIN_PASS = "AAIIBS@2026"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def auth():
    r = requests.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASS})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


# ---------- template
def test_template_requires_auth():
    r = requests.get(f"{API}/admin/ipads/template")
    assert r.status_code == 401


def test_template_download_xlsx(auth):
    r = requests.get(f"{API}/admin/ipads/template", headers=auth)
    assert r.status_code == 200
    assert XLSX_MIME in r.headers.get("content-type", "")
    assert "attachment" in r.headers.get("content-disposition", "").lower()
    # signature must be a valid xlsx (PK zip magic)
    assert r.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ["serial_number", "version", "storage", "purchase_year", "color", "notes"]


# ---------- bulk import
def _make_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_bulk_import_non_excel_rejected(auth):
    files = {"file": ("notes.txt", b"hello", "text/plain")}
    r = requests.post(f"{API}/admin/ipads/bulk", headers=auth, files=files)
    assert r.status_code == 400


def test_bulk_import_created_skipped_errors(auth):
    header = ["serial_number", "version", "storage", "purchase_year", "color", "notes"]
    rows = [
        header,
        ["TESTBULK001", "iPad Gen 10", "256GB", 2026, "Silver", "bulk row 1"],
        ["TESTBULK002", "iPad Gen 9", "64GB", 2025, "Space Gray", "bulk row 2"],
        ["DMPGX1AAAA01", "iPad Gen 10", "256GB", 2026, "Silver", "dup existing"],  # existing seed -> skipped
        ["", "iPad", "64GB", 2024, "", "missing serial -> error"],
    ]
    buf = _make_xlsx(rows)
    files = {"file": ("import.xlsx", buf.getvalue(), XLSX_MIME)}
    r = requests.post(f"{API}/admin/ipads/bulk", headers=auth, files=files)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["created"] == 2
    assert data["skipped"] == 1
    assert isinstance(data["errors"], list) and len(data["errors"]) == 1

    # verify persisted + cleanup
    lst = requests.get(f"{API}/admin/ipads", headers=auth).json()
    by_serial = {x["serial_number"]: x for x in lst}
    assert "TESTBULK001" in by_serial
    assert "TESTBULK002" in by_serial

    # cleanup
    for s in ("TESTBULK001", "TESTBULK002"):
        rid = by_serial[s]["id"]
        d = requests.delete(f"{API}/admin/ipads/{rid}", headers=auth)
        assert d.status_code == 200

    # re-run: now both should be created again (verifies cleanup) - skip to save time
    lst2 = requests.get(f"{API}/admin/ipads", headers=auth).json()
    serials2 = {x["serial_number"] for x in lst2}
    assert "TESTBULK001" not in serials2
    assert "TESTBULK002" not in serials2


def test_seed_data_intact_after_cleanup(auth):
    """Ensure exactly the 5 demo serials still present after bulk cleanup."""
    lst = requests.get(f"{API}/admin/ipads", headers=auth).json()
    serials = {x["serial_number"] for x in lst}
    expected = {"DMPGX1AAAA01", "DMPGX1AAAA02", "F9FZ2BBBB03", "F9FZ2BBBB04", "GHKL3CCCC05"}
    assert expected.issubset(serials)
    # no leftover TEST serials
    assert not any(s.startswith("TEST") for s in serials), f"Leftover TEST serials: {[s for s in serials if s.startswith('TEST')]}"
