from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import logging
import secrets
import string
from datetime import datetime, timezone, timedelta
from typing import List, Optional

import bcrypt
import jwt
from bson import ObjectId
from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, UploadFile, File
from fastapi.responses import Response
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from openpyxl import Workbook, load_workbook
import uuid

from pdf_pakta import build_pakta_pdf

# ---------------------------------------------------------------- setup
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="AAIIBS iPad Lungsuran System")
api = APIRouter(prefix="/api")

JWT_ALG = "HS256"
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("aaiibs")


# ---------------------------------------------------------------- auth utils
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_access_token(uid: str, email: str) -> str:
    payload = {"sub": uid, "email": email, "type": "access",
               "exp": datetime.now(timezone.utc) + timedelta(hours=12)}
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALG)


async def get_current_admin(request: Request) -> dict:
    token = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
    if not token:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALG])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return {"id": str(user["_id"]), "email": user["email"], "name": user.get("name", "Admin")}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ---------------------------------------------------------------- models
class LoginInput(BaseModel):
    email: EmailStr
    password: str


class IpadInput(BaseModel):
    serial_number: str
    version: str
    storage: str
    purchase_year: int
    color: Optional[str] = ""
    notes: Optional[str] = ""


class CodeInput(BaseModel):
    serial_number: str
    target_name: Optional[str] = ""
    count: int = 1


class BatchCodeInput(BaseModel):
    serial_numbers: List[str]
    target_name: Optional[str] = ""
    count: int = 1


class ValidateInput(BaseModel):
    code: str


class PaktaInput(BaseModel):
    code: str
    nama: str
    nik: str
    jabatan: str
    unit: str
    tanggal_peminjaman: str
    signature: str


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def gen_code(n=6):
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(secrets.choice(alphabet) for _ in range(n))


# ---------------------------------------------------------------- auth routes
@api.post("/auth/login")
async def login(body: LoginInput, response: Response):
    email = body.email.lower()
    ident = f"login:{email}"
    att = await db.login_attempts.find_one({"identifier": ident})
    if att and att.get("count", 0) >= 5:
        locked_until = att.get("locked_until")
        if locked_until and datetime.fromisoformat(locked_until) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Terlalu banyak percobaan. Coba lagi nanti.")
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": ident},
            {"$inc": {"count": 1},
             "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}},
            upsert=True,
        )
        raise HTTPException(status_code=401, detail="Email atau password salah")
    await db.login_attempts.delete_one({"identifier": ident})
    token = create_access_token(str(user["_id"]), user["email"])
    response.set_cookie("access_token", token, httponly=True, secure=True,
                        samesite="none", max_age=43200, path="/")
    return {"access_token": token,
            "user": {"id": str(user["_id"]), "email": user["email"], "name": user.get("name", "Admin")}}


@api.get("/auth/me")
async def me(admin: dict = Depends(get_current_admin)):
    return admin


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


# ---------------------------------------------------------------- ipad admin routes
@api.post("/admin/ipads")
async def add_ipad(body: IpadInput, admin: dict = Depends(get_current_admin)):
    serial = body.serial_number.strip().upper()
    if await db.ipads.find_one({"serial_number": serial}):
        raise HTTPException(status_code=400, detail="Serial number sudah terdaftar")
    doc = {
        "id": str(uuid.uuid4()),
        "serial_number": serial,
        "version": body.version.strip(),
        "storage": body.storage.strip(),
        "purchase_year": body.purchase_year,
        "color": (body.color or "").strip(),
        "notes": (body.notes or "").strip(),
        "created_at": now_iso(),
    }
    await db.ipads.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/admin/ipads/{ipad_id}")
async def update_ipad(ipad_id: str, body: IpadInput, admin: dict = Depends(get_current_admin)):
    serial = body.serial_number.strip().upper()
    existing = await db.ipads.find_one({"id": ipad_id})
    if not existing:
        raise HTTPException(status_code=404, detail="iPad tidak ditemukan")
    clash = await db.ipads.find_one({"serial_number": serial, "id": {"$ne": ipad_id}})
    if clash:
        raise HTTPException(status_code=400, detail="Serial number sudah dipakai iPad lain")
    await db.ipads.update_one({"id": ipad_id}, {"$set": {
        "serial_number": serial, "version": body.version.strip(), "storage": body.storage.strip(),
        "purchase_year": body.purchase_year, "color": (body.color or "").strip(),
        "notes": (body.notes or "").strip(),
    }})
    return {"ok": True}


@api.delete("/admin/ipads/{ipad_id}")
async def delete_ipad(ipad_id: str, admin: dict = Depends(get_current_admin)):
    ipad = await db.ipads.find_one({"id": ipad_id})
    if not ipad:
        raise HTTPException(status_code=404, detail="iPad tidak ditemukan")
    if await db.paktas.find_one({"serial_number": ipad["serial_number"]}):
        raise HTTPException(status_code=400, detail="Tidak bisa dihapus, sudah ada riwayat pakta.")
    await db.ipads.delete_one({"id": ipad_id})
    await db.codes.delete_many({"serial_number": ipad["serial_number"], "status": "active"})
    return {"ok": True}


@api.get("/admin/ipads")
async def admin_list_ipads(admin: dict = Depends(get_current_admin)):
    return await _ipads_with_holders()


IMPORT_HEADERS = ["serial_number", "version", "storage", "purchase_year", "color", "notes"]


@api.get("/admin/ipads/template")
async def ipad_template(admin: dict = Depends(get_current_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "iPad"
    ws.append(IMPORT_HEADERS)
    ws.append(["DMPGX1AAAA01", "iPad Gen 10", "256GB", 2026, "Silver", "unit baru"])
    ws.append(["F9FZ2BBBB03", "iPad Gen 7", "32GB", 2021, "Space Gray", "lungsuran"])
    for i, col in enumerate(["A", "B", "C", "D", "E", "F"]):
        ws.column_dimensions[col].width = [18, 16, 12, 14, 14, 24][i]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_import_ipad.xlsx"'},
    )


@api.post("/admin/ipads/bulk")
async def bulk_import_ipads(file: UploadFile = File(...), admin: dict = Depends(get_current_admin)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File harus berformat Excel (.xlsx)")
    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Gagal membaca file Excel")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    created, skipped, errors = 0, 0, []
    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = list(row) + [None] * (6 - len(row))
        serial = str(cells[0]).strip().upper() if cells[0] else ""
        version = str(cells[1]).strip() if cells[1] else ""
        storage = str(cells[2]).strip() if cells[2] else ""
        try:
            year = int(float(cells[3])) if cells[3] not in (None, "") else 0
        except (ValueError, TypeError):
            year = 0
        color = str(cells[4]).strip() if cells[4] else ""
        notes = str(cells[5]).strip() if cells[5] else ""
        if not serial or not version or not storage or not year:
            errors.append(f"Baris {idx}: serial/versi/penyimpanan/tahun wajib diisi")
            continue
        if await db.ipads.find_one({"serial_number": serial}):
            skipped += 1
            continue
        await db.ipads.insert_one({
            "id": str(uuid.uuid4()), "serial_number": serial, "version": version,
            "storage": storage, "purchase_year": year, "color": color,
            "notes": notes, "created_at": now_iso(),
        })
        created += 1
    return {"created": created, "skipped": skipped, "errors": errors[:20]}


# ---------------------------------------------------------------- code admin routes
@api.post("/admin/codes")
async def create_codes(body: CodeInput, admin: dict = Depends(get_current_admin)):
    ipad = await db.ipads.find_one({"serial_number": body.serial_number.strip().upper()})
    if not ipad:
        raise HTTPException(status_code=404, detail="iPad tidak ditemukan")
    return await _gen_codes_for_ipad(ipad, body.count, body.target_name)


async def _gen_codes_for_ipad(ipad, count, target_name):
    count = max(1, min(int(count), 20))
    created = []
    for _ in range(count):
        code = gen_code()
        while await db.codes.find_one({"code": code}):
            code = gen_code()
        doc = {
            "id": str(uuid.uuid4()), "code": code,
            "serial_number": ipad["serial_number"], "version": ipad["version"],
            "storage": ipad["storage"], "purchase_year": ipad["purchase_year"],
            "target_name": (target_name or "").strip(),
            "status": "active", "used_by": None,
            "created_at": now_iso(), "used_at": None,
        }
        await db.codes.insert_one(doc)
        doc.pop("_id", None)
        created.append(doc)
    return created


@api.post("/admin/codes/batch")
async def create_codes_batch(body: BatchCodeInput, admin: dict = Depends(get_current_admin)):
    serials = list({s.strip().upper() for s in body.serial_numbers if s and s.strip()})
    if not serials:
        raise HTTPException(status_code=400, detail="Pilih minimal satu iPad")
    created, not_found = [], []
    for serial in serials:
        ipad = await db.ipads.find_one({"serial_number": serial})
        if not ipad:
            not_found.append(serial)
            continue
        created.extend(await _gen_codes_for_ipad(ipad, body.count, body.target_name))
    return {"created": created, "ipad_count": len(serials) - len(not_found),
            "code_count": len(created), "not_found": not_found}


@api.get("/admin/codes")
async def admin_list_codes(admin: dict = Depends(get_current_admin)):
    codes = await db.codes.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return codes


@api.delete("/admin/codes/{code_id}")
async def delete_code(code_id: str, admin: dict = Depends(get_current_admin)):
    code = await db.codes.find_one({"id": code_id})
    if not code:
        raise HTTPException(status_code=404, detail="Kode tidak ditemukan")
    if code["status"] == "used":
        raise HTTPException(status_code=400, detail="Kode sudah dipakai, tidak bisa dihapus")
    await db.codes.delete_one({"id": code_id})
    return {"ok": True}


@api.get("/admin/paktas")
async def admin_list_paktas(admin: dict = Depends(get_current_admin)):
    return await db.paktas.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


# ---------------------------------------------------------------- public: token + pakta
@api.post("/codes/validate")
async def validate_code(body: ValidateInput):
    code = body.code.strip().upper()
    doc = await db.codes.find_one({"code": code}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Kode tidak valid")
    if doc["status"] == "used":
        raise HTTPException(status_code=400, detail="Kode ini sudah digunakan untuk mengisi pakta")
    return {
        "code": doc["code"], "serial_number": doc["serial_number"],
        "version": doc["version"], "storage": doc["storage"],
        "purchase_year": doc["purchase_year"], "target_name": doc.get("target_name", ""),
    }


@api.post("/pakta/submit")
async def submit_pakta(body: PaktaInput):
    code = body.code.strip().upper()
    codedoc = await db.codes.find_one({"code": code})
    if not codedoc:
        raise HTTPException(status_code=404, detail="Kode tidak valid")
    if codedoc["status"] == "used":
        raise HTTPException(status_code=400, detail="Kode ini sudah digunakan")
    if not body.signature or not body.signature.startswith("data:image") or "," not in body.signature:
        raise HTTPException(status_code=400, detail="Tanda tangan wajib diisi")

    prior = await db.paktas.count_documents({"serial_number": codedoc["serial_number"]})
    pakta = {
        "id": str(uuid.uuid4()),
        "serial_number": codedoc["serial_number"],
        "ipad_version": codedoc["version"],
        "storage": codedoc["storage"],
        "purchase_year": codedoc["purchase_year"],
        "nama": body.nama.strip(),
        "nik": body.nik.strip(),
        "jabatan": body.jabatan.strip(),
        "unit": body.unit.strip(),
        "tanggal_peminjaman": body.tanggal_peminjaman,
        "tanggal_pengisian": now_iso(),
        "signature": body.signature,
        "nama_terang": body.nama.strip(),
        "code": code,
        "sequence": prior + 1,
        "created_at": now_iso(),
    }
    await db.paktas.insert_one(pakta)
    await db.codes.update_one({"code": code},
                              {"$set": {"status": "used", "used_by": pakta["id"], "used_at": now_iso()}})
    return {"id": pakta["id"]}


@api.get("/pakta/{pakta_id}")
async def get_pakta(pakta_id: str):
    doc = await db.paktas.find_one({"id": pakta_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Pakta tidak ditemukan")
    return doc


@api.get("/pakta/{pakta_id}/pdf")
async def pakta_pdf(pakta_id: str):
    doc = await db.paktas.find_one({"id": pakta_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Pakta tidak ditemukan")
    pdf = build_pakta_pdf(doc)
    safe = doc["nama"].replace(" ", "_")
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="Pakta_{safe}.pdf"'})


# ---------------------------------------------------------------- public dashboard / trace
async def _ipads_with_holders():
    ipads = await db.ipads.find({}, {"_id": 0}).sort("purchase_year", -1).to_list(1000)
    out = []
    for ip in ipads:
        chain = await db.paktas.find(
            {"serial_number": ip["serial_number"]}, {"_id": 0, "signature": 0}
        ).sort("created_at", 1).to_list(100)
        current = chain[-1] if chain else None
        out.append({
            **ip,
            "current_holder": current["nama"] if current else None,
            "current_holder_unit": (f"{current['jabatan']} / {current['unit']}" if current else None),
            "holder_count": len(chain),
            "is_lungsuran": len(chain) > 1,
            "chain": chain,
        })
    return out


@api.get("/public/ipads")
async def public_ipads():
    return await _ipads_with_holders()


@api.get("/public/trace/{serial}")
async def public_trace(serial: str):
    serial = serial.strip().upper()
    ipad = await db.ipads.find_one({"serial_number": serial}, {"_id": 0})
    if not ipad:
        raise HTTPException(status_code=404, detail="Serial number tidak ditemukan")
    chain = await db.paktas.find(
        {"serial_number": serial}, {"_id": 0, "signature": 0}
    ).sort("created_at", 1).to_list(100)
    return {"ipad": ipad, "chain": chain}


@api.get("/public/stats")
async def public_stats():
    ipads = await db.ipads.find({}, {"_id": 0}).to_list(1000)
    paktas = await db.paktas.find({}, {"_id": 0, "signature": 0}).to_list(2000)

    by_version, by_storage, by_year = {}, {}, {}
    for ip in ipads:
        by_version[ip["version"]] = by_version.get(ip["version"], 0) + 1
        by_storage[ip["storage"]] = by_storage.get(ip["storage"], 0) + 1
        y = str(ip["purchase_year"])
        by_year[y] = by_year.get(y, 0) + 1

    serials_with_holder = {p["serial_number"] for p in paktas}
    chain_counts = {}
    for p in paktas:
        chain_counts[p["serial_number"]] = chain_counts.get(p["serial_number"], 0) + 1
    lungsuran_count = sum(1 for v in chain_counts.values() if v > 1)

    recent = sorted(paktas, key=lambda x: x.get("created_at", ""), reverse=True)[:8]
    recent = [{"nama": r["nama"], "unit": r["unit"], "jabatan": r["jabatan"],
               "serial_number": r["serial_number"], "ipad_version": r["ipad_version"],
               "storage": r["storage"], "created_at": r["created_at"], "id": r["id"]} for r in recent]

    active_codes = await db.codes.count_documents({"status": "active"})

    return {
        "total_ipads": len(ipads),
        "total_paktas": len(paktas),
        "active_holders": len(serials_with_holder),
        "available_ipads": len(ipads) - len(serials_with_holder),
        "lungsuran_count": lungsuran_count,
        "active_codes": active_codes,
        "by_version": [{"name": k, "value": v} for k, v in sorted(by_version.items())],
        "by_storage": [{"name": k, "value": v} for k, v in sorted(by_storage.items())],
        "by_year": [{"name": k, "value": v} for k, v in sorted(by_year.items())],
        "recent": recent,
    }


@api.get("/")
async def root():
    return {"message": "AAIIBS iPad Lungsuran System API"}


# ---------------------------------------------------------------- startup
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.ipads.create_index("serial_number", unique=True)
    await db.codes.create_index("code", unique=True)
    await db.paktas.create_index("serial_number")
    admin_email = os.environ["ADMIN_EMAIL"].lower()
    admin_pw = os.environ["ADMIN_PASSWORD"]
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "email": admin_email, "password_hash": hash_password(admin_pw),
            "name": "Administrator", "role": "admin", "created_at": now_iso(),
        })
        logger.info("Seeded admin user %s", admin_email)
    elif not verify_password(admin_pw, existing["password_hash"]):
        await db.users.update_one({"email": admin_email},
                                  {"$set": {"password_hash": hash_password(admin_pw)}})


@app.on_event("shutdown")
async def shutdown():
    client.close()


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)
